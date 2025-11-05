import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from collections import defaultdict
from typing import List, Dict, Tuple
import pandas as pd
# --- MODIFIED IMPORT ---
from models import Place, SquareGrid # Now import SquareGrid
from config import COMBO, NUM_CELLS, GAMMAS, DATASET_NAMES

# --- MODIFIED IMPORT ---
from baseline_iadu import load_dataset, plot_selected
from grid_iadu import grid_iadu
from extension_sampling import grid_sampling 
from biased_sampling import biased_sampling
# ---

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- NEW IMPORTS for PLOTTING ---
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
# ---

# Updated experiment name for the focused run
EXPERIMENT_NAME = "grid_VS_extension_VS_biased"
SHAPES = DATASET_NAMES

def run_experiment():
    """
    Runs a simplified experiment comparing:
    1. Grid IAdU (Main Paper)
    2. Grid Sampling (Extension Paper, Option 3)
    3. Biased Sampling (From hardcore_exp.py)
    
    Logs H(R), psS, psR, and lenCL to Excel.
    Generates a new 'experiment_plots.pdf' with visual results *including grid lines*.
    """
    log = defaultdict(list)

    # --- NEW: Initialize PDF file ---
    pdf_path = f"{EXPERIMENT_NAME}_plots.pdf"
    pdf_pages = PdfPages(pdf_path)
    # ---

    for (K, k) in COMBO:
        for g in GAMMAS:
            W = K / (g * k)
            print(f"Comparing Grid_IAdU vs. Grid_Sampling vs. Biased_Sampling | K={K}, k={k}, Gs={NUM_CELLS}")

            for shape in SHAPES:
                for G in NUM_CELLS:
                    print(f"  Shape={shape}, G={G}")
                    
                    S: List[Place] = load_dataset(shape, K)
                    
                    if not S:
                        print(f"    Skipping shape {shape} for K={K}, no data loaded.")
                        continue

                    # --- NEW: Create the grid object ---
                    # This is the same grid that grid_iadu and grid_sampling will use
                    try:
                        grid = SquareGrid(S, G)
                    except ValueError:
                        print(f"    Skipping shape {shape}, K={K}. Invalid data for grid.")
                        continue
                    # ---

                    # --- 1. Run Grid IAdU (Main Paper) ---
                    R_grid, score_grid, grid_pss_sum, grid_psr_sum, _, _, lenCL = grid_iadu(S, k, W, G)

                    # --- 2. Run Grid Sampling (Extension) ---
                    R_grid_samp, score_grid_samp, grid_samp_pss_sum, grid_samp_psr_sum, _, _, _ = grid_sampling(S, k, W, G)

                    # --- 3. Run Biased Sampling (FIXED UNPACK) ---
                    R_biased, score_biased, biased_pss_sum, biased_psr_sum, _ = biased_sampling(S, k, W)

                    # --- 4. Log all comparison data ---
                    log_entry = {
                        "shape": shape,
                        "K": K,
                        "k": k,
                        "W": W,
                        "G": G,
                        "lenCL": lenCL, 
                        "K/(k*g)": f"K/(k * {g})",

                        # Grid IAdU Results
                        "grid_iadu_hpfr": score_grid,
                        "grid_iadu_pss_sum": grid_pss_sum,
                        "grid_iadu_psr_sum": grid_psr_sum,
                        
                        # Grid Sampling Results
                        "grid_sampling_hpfr": score_grid_samp,
                        "grid_sampling_pss_sum": grid_samp_pss_sum,
                        "grid_sampling_psr_sum": grid_samp_psr_sum,

                        # Biased Sampling Results
                        "biased_hpfr": score_biased,
                        "biased_pss_sum": biased_pss_sum,
                        "biased_psr_sum": biased_psr_sum,
                    }
                    log[(K, k, g, G)].append(log_entry)

                    # --- UPDATED: Plotting logic ---
                    fig, axes = plt.subplots(1, 3, figsize=(21, 7))
                    
                    fig.suptitle(f"Shape: {shape}  |  K={K}, k={k}  |  G={G} (Ax={grid.Ax}, Ay={grid.Ay})  |  lenCL={lenCL}", fontsize=16)
                    
                    # Plot 1: Grid IAdU (now passing grid=grid)
                    plot_selected(S, R_grid, f"Grid IAdU\nHPFR: {score_grid: .4f}", axes[0], grid=grid)
                    
                    # Plot 2: Grid Sampling (now passing grid=grid)
                    plot_selected(S, R_grid_samp, f"Grid Sampling (Extension)\nHPFR: {score_grid_samp: .4f}", axes[1], grid=grid)
                    
                    # Plot 3: Biased Sampling (now passing grid=grid)
                    plot_selected(S, R_biased, f"Biased Sampling\nHPFR: {score_biased: .4f}", axes[2], grid=grid)
                    
                    pdf_pages.savefig(fig)
                    plt.close(fig) 
                    # --- End of updated plotting logic ---

    # --- NEW: Close the PDF file after the loop ---
    pdf_pages.close()
    print(f"\nSuccessfully saved plots to {pdf_path}")
    # ---

    avg_log = compute_average_log(log)
    save_outputs(avg_log)

    
def save_outputs(log: Dict):
    """
    Saves a simplified Excel file with the direct comparison.
    """
    def smart_round(value):
        if value is None:
            return None
        if value == 0:
            return 0.0
        elif abs(value) >= 0.01:
            return round(value, 3)
        else:
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"

    all_rows = []
    for row in log.values():
        for k, v in list(row.items()):
            if isinstance(v, float):
                row[k] = smart_round(v)
            if k == "lenCL" and isinstance(v, float):
                 row[k] = round(v, 1)
        all_rows.append(row)

    df = pd.DataFrame(all_rows)

    if "shape" in df.columns:
        df.drop(columns=["shape"], inplace=True)

    sort_cols = [c for c in ["K", "k", "K/(k*g)", "G"] if c in df.columns]
    if sort_cols:
        df.sort_values(by=sort_cols, ascending=[True] * len(sort_cols), inplace=True)

    setup_cols = ["K", "k", "W", "K/(k*g)", "G", "lenCL"]
    
    score_cols = [
        "grid_iadu_hpfr",
        "grid_iadu_pss_sum",
        "grid_iadu_psr_sum",
        
        "grid_sampling_hpfr",
        "grid_sampling_pss_sum",
        "grid_sampling_psr_sum",
        
        "biased_hpfr",
        "biased_pss_sum",
        "biased_psr_sum",
    ]
    
    all_cols = setup_cols + score_cols

    for col in all_cols:
        if col not in df.columns:
            df[col] = None
    df = df[all_cols]

    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # === Styling ===
    wb = load_workbook(xlsx_name)
    ws = wb.active
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    group_fills = {
        "setup": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "score": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }

    def apply_group_style(cols, fill, border_after=False):
        for i, col_name in enumerate(cols):
            if col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    cell.fill = fill
                    current_border = cell.border
                    if border_after and i == len(cols) - 1:
                        cell.border = Border(left=current_border.left, right=Side(style='thick'),
                                              top=current_border.top, bottom=current_border.bottom)
                    else:
                        cell.border = thin_border

    apply_group_style(setup_cols, group_fills["setup"], border_after=True)
    apply_group_style(score_cols, group_fills["score"], border_after=False)

    # autosize
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(xlsx_name)
    print(f"Results saved to {xlsx_name}")

def compute_average_log(
    log: Dict[Tuple[int, int, int, int], List[Dict]]
) -> Dict[Tuple[int, int, int, int], Dict]:
    """
    Averages results per (K, k, g, G) key, across all shapes.
    """
    avg_log: Dict[Tuple[int, int, int, int], Dict] = {}

    for key, rows in log.items():  # key = (K, k, g, G)
        if not rows:
            continue
        
        # Setup the output dictionary with non-numeric fields
        out = {
            "K": key[0],
            "k": key[1],
            "W": rows[0]["W"],
            "K/(k*g)": rows[0]["K/(k*g)"],
            "G": key[3],
        }
        
        all_fields = set().union(*[r.keys() for r in rows])
        for fname in all_fields:
            if fname in {"shape", "K", "k", "g", "G", "W", "K/(k*g)"}:
                continue
            
            vals = [r[fname] for r in rows if isinstance(r.get(fname), (int, float))]
            if vals:
                out[fname] = sum(vals) / len(vals)
            elif fname not in out:
                out[fname] = None 

        avg_log[key] = out

    return avg_log
                        
if __name__ == "__main__":
    run_experiment()