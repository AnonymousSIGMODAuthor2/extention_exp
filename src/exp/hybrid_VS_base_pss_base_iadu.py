import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "src")))

from collections import defaultdict
from typing import List, Dict
import pandas as pd
from models import Place
from config import COMBO, NUM_CELLS, SIMULATED_DATASETS
from baseline_iadu import load_dataset, iadu
from hybrid_sampling import hybrid

EXPERIMENT_NAME = "hybrid_VS_base_psS_base_iadu"
SHAPES = SIMULATED_DATASETS  # or DATASET_NAMES for real datasets


def run_experiment():
    log = defaultdict(list)

    for (K, k) in COMBO:
        W = K / k
        print(f"Comparing Hybrid vs Baseline on HPFR | K={K}, k={k}, W={W:.2f}")

        for shape in SHAPES:
            for G in NUM_CELLS:
                print(f"  Shape={shape}, G={G}")
                try:
                    S: List[Place] = load_dataset(shape, K, k, G)
                except FileNotFoundError:
                    print(f"    ✗ Missing dataset: {shape}_K{K}_k{k}_G{G}.pkl")
                    continue

                # === Baseline IAdU (ground truth) ===
                R_base, psS_base, score_base, t_base_prep, t_base_select = iadu(S, k, W)


                # === Hybrid Sampling using same psS_base ===
                R_hybrid, score_hybrid, t_hybrid_prep, t_hybrid_select, W_hybrid, K_sample = hybrid(S, k, W)

                # Log metrics
                log[(K, k, G)].append({
                    "shape": shape,
                    "K": K,
                    "k": k,
                    "W": W,
                    "K'": K_sample,
                    "W'": W_hybrid,
                    "baseline_score": score_base,
                    "hybrid_score": score_hybrid,
                    "hpfr_error%": 100 * abs(score_hybrid - score_base) / score_base,
                    "baseline_pss_time": t_base_prep,
                    "baseline_iadu_time": t_base_select,
                    "hybrid_pss_time": t_hybrid_prep,
                    "hybrid_iadu_time": t_hybrid_select,
                    "baseline_total_time": t_base_prep + t_base_select,
                    "hybrid_total_time": t_hybrid_prep + t_hybrid_select,
                    "speedup_total": (t_base_prep + t_base_select) / (t_hybrid_prep + t_hybrid_select),
                })

    final_log = []
    for (K, k, G), rows in log.items():
        count = len(rows)
        accum = {}
        for row in rows:
            for key, val in row.items():
                if isinstance(val, (int, float)):
                    accum[key] = accum.get(key, 0) + val

        avg_row = {key: accum[key] / count for key in accum}
        final_log.append(avg_row)

    save_outputs(final_log)


def save_outputs(log: List[Dict]):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def smart_round(value):
        if value == 0:
            return 0.0
        elif abs(value) >= 0.01:
            return round(value, 3)
        else:
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"

    for row in log:
        for k, v in row.items():
            if isinstance(v, float):
                row[k] = smart_round(v)

    df = pd.DataFrame(log)

    # === Define column blocks ===
    setup_cols = ["K", "k", "K'", "W", "W'"]
    score_cols = ["baseline_score", "hybrid_score", "hpfr_error%"]
    prep_cols = ["baseline_pss_time", "hybrid_pss_time"]
    select_cols = ["baseline_iadu_time", "hybrid_iadu_time"]
    total_cols = ["baseline_total_time", "hybrid_total_time", "speedup_total"]

    all_cols = setup_cols + score_cols + prep_cols + select_cols + total_cols
    remaining_cols = [col for col in df.columns if col not in all_cols]
    df = df[all_cols + remaining_cols]

    # Save raw Excel first
    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # === Excel Styling ===
    wb = load_workbook(xlsx_name)
    ws = wb.active

    fills = {
        "setup": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "scores": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "pss": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "iadu": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "total": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    }

    block_map = {}
    for col in setup_cols:
        block_map[col] = "setup"
    for col in score_cols:
        block_map[col] = "scores"
    for col in prep_cols:
        block_map[col] = "pss"
    for col in select_cols:
        block_map[col] = "iadu"
    for col in total_cols:
        block_map[col] = "total"

    thin = Side(style="thin")
    thick = Side(style="thick")
    center_align = Alignment(horizontal="center")

    headers = [cell.value for cell in ws[1]]

    # Determine block edges
    block_ends = []
    for i, h in enumerate(headers):
        if h in score_cols + prep_cols + select_cols + total_cols:
            block_ends.append(i)

    def style_header_row(row_cells):
        for j, cell in enumerate(row_cells):
            block = block_map.get(cell.value, None)
            if block:
                cell.fill = fills[block]
            cell.font = Font(bold=True)
            cell.alignment = center_align

            if j in block_ends:
                cell.border = Border(top=thin, bottom=thin, right=thick)
            elif j == 0 or j - 1 in block_ends:
                cell.border = Border(top=thin, bottom=thin, left=thick)
            else:
                cell.border = Border(top=thin, bottom=thin)

    def style_data_row(row):
        for j, cell in enumerate(row):
            header = headers[j]
            block = block_map.get(header, None)
            if block:
                cell.fill = fills[block]
            cell.alignment = center_align

            if j in block_ends:
                cell.border = Border(bottom=thin, right=thick)
            elif j == 0 or j - 1 in block_ends:
                cell.border = Border(bottom=thin, left=thick)
            else:
                cell.border = Border(bottom=thin)

    # Style header
    style_header_row(ws[1])

    # Style all data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        style_data_row(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(xlsx_name)
    print(f"Styled and saved: {xlsx_name}")

if __name__ == "__main__":
    run_experiment()
