import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from biased_sampling import biased_sampling
from collections import defaultdict
from typing import List, Dict, Tuple
import pandas as pd
from models import Place, SquareGrid
from config import COMBO, NUM_CELLS, GAMMAS, DATASET_NAMES
from baseline_iadu import base_precompute, iadu, load_dataset
from hybrid_sampling import hybrid, hybrid_on_grid
from grid_iadu import base_iadu_on_grid, grid_iadu, virtual_grid_based_algorithm
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPERIMENT_NAME = "pss_error_comparison"
SHAPES = DATASET_NAMES

def run_experiment():
    log = defaultdict(list)

    for (K, k) in COMBO:
        for g in GAMMAS:
            W = K / (g * k)
            print(f"Comparing ALL methods on HPFR and times | K={K}, k={k}, g={g}, W={W:.2f}")

            for shape in SHAPES:
                for G in NUM_CELLS:
                    print(f"  Shape={shape}, G={G}")
                    
                    S: List[Place] = load_dataset(shape, K)
                    

                    # --- use three K' values ---
                    K_samples = [int(K * 0.2)]

                    # Baselines / Grid / Biased (do once per dataset)
                    exact_psS, exact_sS, prep_time = base_precompute(S)
                    base_pss_sum = sum(exact_psS[p.id] for p in S)
                    
                    grid = SquareGrid(S, G)
                    CL = grid.get_full_cells()
                    psS, sS , prep_time = virtual_grid_based_algorithm(CL,S)
                    grid_pss_sum = sum(psS[p.id] for p in S)
                    
                    approx_error = sum(
                    100 * abs(psS[p.id] - exact_psS[p.id]) / exact_psS[p.id] for p in S
                    ) / len(S)


                    log[(K, k, g, (G, len(CL)))].append({
                            "shape": shape,
                            "K": K,
                            "k": k,
                            "W": W,
                            "K/(k*g)": f"K/(k * {g})",

                            "baseline_psS_sum": base_pss_sum,
                            
                            "grid_psS_sum": grid_pss_sum,

                            "grid_error": approx_error,
                    })

    avg_log = compute_average_log(log)
    save_outputs(avg_log)


def save_outputs(log: Dict):
    def smart_round(value):
        if value is None:
            return None
        if value == 0:
            return 0.0
        if isinstance(value, (int, float)):
            if abs(value) >= 0.01:
                return round(value, 3)
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"
        return value

    all_rows = []
    for row in log.values():
        r = dict(row)
        for k, v in list(r.items()):
            if isinstance(v, (int, float)) or v is None:
                r[k] = smart_round(v)
        all_rows.append(r)

    df = pd.DataFrame(all_rows)

    if "shape" in df.columns:
        df.drop(columns=["shape"], inplace=True)

    sort_cols = [c for c in ["K", "k", "K/(k*g)", "G", "|CL|"] if c in df.columns]
    if sort_cols:
        df.sort_values(by=sort_cols, ascending=[True] * len(sort_cols), inplace=True)

    setup_cols = ["K", "k", "W", "K/(k*g)", "G", "|CL|"]
    pss_cols = [
        "grid_error",
        "baseline_psS_sum",
        "grid_psS_sum",
    ]
    all_cols = setup_cols + pss_cols
    for col in all_cols:
        if col not in df.columns:
            df[col] = None
    df = df[all_cols]

    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # --- styling (unchanged) ---
    wb = load_workbook(xlsx_name)
    ws = wb.active
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    group_fills = {
        "setup": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "pss": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }

    def apply_group_style(cols, fill, border_after=False):
        for i, col_name in enumerate(cols):
            if col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    cell.fill = fill
                    if border_after and i == len(cols) - 1:
                        cell.border = Border(right=Side(style='thick'))
                    else:
                        cell.border = thin_border

    apply_group_style(setup_cols, group_fills["setup"], border_after=True)
    apply_group_style(pss_cols, group_fills["pss"], border_after=True)

    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(xlsx_name)
    print(f"Results saved to {xlsx_name}")

from collections import defaultdict
from typing import Dict, Tuple, List

def compute_average_log(
    log: Dict[Tuple[int, int, int, Tuple[int, int]], List[Dict]]
) -> Dict[Tuple[int, int, int, int], Dict]:
    """
    Re-group by (K, k, g, G) and average numeric fields across shapes, including |CL|.
    G stays fixed per row; |CL| is averaged.
    """
    # 1) Rebucket: (K,k,g,(G,|CL|))  ->  (K,k,g,G) with rows carrying |CL| inside
    buckets: Dict[Tuple[int, int, int, int], List[Dict]] = defaultdict(list)

    for key, rows in log.items():
        if not rows:
            continue

        K_val, k_val, g_val, g_cl = key
        if isinstance(g_cl, tuple) and len(g_cl) == 2:
            G_val, cl_len = g_cl
        else:
            G_val, cl_len = g_cl, None

        for r in rows:
            r2 = dict(r)
            # Inject |CL| from the key so it can be averaged like other numeric fields
            r2["|CL|"] = cl_len
            buckets[(K_val, k_val, g_val, G_val)].append(r2)

    # 2) Average within each (K,k,g,G)
    avg_log: Dict[Tuple[int, int, int, int], Dict] = {}

    for agg_key, rows in buckets.items():
        if not rows:
            continue

        K_val, k_val, g_val, G_val = agg_key
        seed = rows[0]

        out = {
            "K": K_val,
            "k": k_val,
            "W": seed.get("W"),              # constant for fixed (K,k,g)
            "K/(k*g)": seed.get("K/(k*g)"),  # you log a display string here
            "G": G_val,
        }

        # union of all fields present
        all_fields = set().union(*[r.keys() for r in rows])

        # Average numeric; keep first (or unanimous) for non-numeric
        # NOTE: do NOT skip "|CL|" — we want it averaged.
        skip = {"shape", "K", "k", "g", "G", "W", "K/(k*g)"}

        for fname in all_fields:
            if fname in skip:
                continue

            vals = [r.get(fname) for r in rows]
            num_vals = [v for v in vals if isinstance(v, (int, float))]

            if num_vals:
                out[fname] = sum(num_vals) / len(num_vals)
            else:
                non_none = [v for v in vals if v is not None]
                if non_none and all(v == non_none[0] for v in non_none):
                    out[fname] = non_none[0]
                else:
                    out[fname] = non_none[0] if non_none else None

        avg_log[(K_val, k_val, g_val, G_val)] = out

    return avg_log

def pct_diff(new, exact):
    return 100 * (new - exact) / exact if exact != 0 else None
                        
if __name__ == "__main__":
    run_experiment()

from collections import defaultdict

